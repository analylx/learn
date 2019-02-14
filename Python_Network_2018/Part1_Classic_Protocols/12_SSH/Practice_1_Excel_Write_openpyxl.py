#!/usr/bin/env python3
# -*- coding=utf-8 -*-
# 本脚由亁颐堂现任明教教主编写，用于乾颐盾Python课程！
# 教主QQ:605658506
# 亁颐堂官网www.qytang.com
# 教主技术进化论拓展你的技术新边疆
# https://ke.qq.com/course/271956?tuin=24199d8a


from openpyxl import Workbook
from openpyxl import load_workbook

dict_excel = {'test123': ('cisco123', 15), 'test456': ('cisco456', 1), 'test789': ('cisco789', 1)}


def excel_write(file='write_pyxl.xlsx', sheel_name='Sheet1', write_dict=dict_excel):
    wb = Workbook()  # 创建xlsx
    ws = wb.create_sheet()  # 创建sheet
    ws.title = sheel_name  # 命名sheet
    # 写入第一行内容
    ws['A1'] = '用户'
    ws['B1'] = '密码'
    ws['C1'] = '级别'
    row_location = 2  # 从第二行开始写入内容
    for x, y in write_dict.items():
        user_locatin = 'A' + str(row_location)
        pass_locatin = 'B' + str(row_location)
        priv_locatin = 'C' + str(row_location)
        ws[user_locatin] = x  # 写入用户
        ws[pass_locatin] = y[0]  # 写入密码
        ws[priv_locatin] = y[1]  # 写入级别
        row_location += 1
    wb.save(file)  # 保存xlsx文件


if __name__ == "__main__":
    excel_write()

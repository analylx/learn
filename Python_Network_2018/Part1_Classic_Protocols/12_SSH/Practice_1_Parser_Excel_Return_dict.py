#!/usr/bin/env python3
# -*- coding=utf-8 -*-
# 本脚由亁颐堂现任明教教主编写，用于乾颐盾Python课程！
# 教主QQ:605658506
# 亁颐堂官网www.qytang.com
# 教主技术进化论拓展你的技术新边疆
# https://ke.qq.com/course/271956?tuin=24199d8a

from openpyxl import Workbook
from openpyxl import load_workbook


def excel_parser_return_dict(file='test.xlsx', sheel_name='Sheet1'):
    data = load_workbook(file)  # 读取xlsx文件
    table = data[sheel_name]  # 读取sheet数据
    excel_dict = {}
    row_location = 0
    for row in table.iter_rows():
        if row_location == 0:  # 跳过第一行！
            row_location += 1
            continue
        else:
            cell_location = 0
            for cell in row:
                if cell_location == 0:  # 读取第一列的用户名
                    tmp_user = cell.value
                    cell_location += 1
                elif cell_location == 1:  # 读取第二列的密码
                    tmp_pass = cell.value
                    cell_location += 1
                elif cell_location == 2:  # 读取第三列的级别
                    tmp_priv = cell.value
                    cell_location += 1
            excel_dict[tmp_user] = tmp_pass, tmp_priv  # 写入字典
        row_location += 1
    return excel_dict  # 返回字典


if __name__ == "__main__":
    print(excel_parser_return_dict('Practice_1_Read_Accounts.xlsx'))
